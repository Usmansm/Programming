#!/usr/bin/env python
# -*- coding: utf-8 -*-
import argparse
from lxml import etree
import openpyxl
import re

namespaces = {'taxii': "http://taxii.mitre.org/messages/taxii_xml_binding-1",
              'taxii_11': "http://taxii.mitre.org/messages/taxii_xml_binding-1.1",
              'tdq': "http://taxii.mitre.org/query/taxii_default_query-1",
              'cyboxCommon': "http://cybox.mitre.org/common-2",
              'cybox': "http://cybox.mitre.org/cybox-2",
              'cyboxVocabs': "http://cybox.mitre.org/default_vocabularies-2",
              'marking': "http://data-marking.mitre.org/Marking-1",
              'simpleMarking': "http://data-marking.mitre.org/extensions/MarkingStructure#Simple-1",
              'tlpMarking': "http://data-marking.mitre.org/extensions/MarkingStructure#TLP-1",
              'TOUMarking': "http://data-marking.mitre.org/extensions/MarkingStructure#Terms_Of_Use-1",
              'opensource': "http://hailataxii.com",
              'edge': "http://soltra.com/",
              'indicator': "http://stix.mitre.org/Indicator-2",
              'ttp': "http://stix.mitre.org/TTP-1",
              'stixCommon': "http://stix.mitre.org/common-1",
              'stixVocabs': "http://stix.mitre.org/default_vocabularies-1",
              'stix': "http://stix.mitre.org/stix-1",
              'xsi': "http://www.w3.org/2001/XMLSchema-instance",
              }


def main(feeds):
    print '-' * 40
    for feed in feeds:
        print 'Start parsing feed `{}`'.format(feed)
        book_name = '{}.xls'.format('.'.join(feed.split('.')[:-1]))
        workbook = openpyxl.Workbook(optimized_write=True)
        ws = workbook.create_sheet()
        ws.title = "Indicators"
        ws.append(('Indicator_id', 'Timestamp', 'Title', 'Type(s)', 'Description', 'Observable', 'Indicated_TTP',
                   'Producer', 'Produced Time', 'Received Time'))

        ws2 = workbook.create_sheet()
        ws2.title = 'TTPs'
        ws2.append(('TTP_Id', 'Timestamp', 'TTP title', 'Type(s)', 'Name(s)', 'Short Description', 'Description'))

        ws3 = workbook.create_sheet()
        ws3.title = 'Observables'
        ws3.append(('Id', 'Title', 'Description', 'Properties'))

        context = etree.iterparse(feed, events=('end',), tag='{%s}Indicator' % namespaces['stix'])
        for event, indicator in context:
            indicator_id = indicator.attrib['id']
            indicator_ts = indicator.attrib['timestamp']
            indicator_title_element = indicator.find('.//indicator:Title', namespaces=namespaces)
            indicator_title = indicator_title_element.text if indicator_title_element is not None else ''
            indicator_type_element = indicator.findall('.//indicator:Type', namespaces=namespaces)
            indicator_types = ','.join([x.text for x in indicator_type_element]) if indicator_type_element else ''
            indicator_description_element = indicator.find('.//indicator:Description', namespaces=namespaces)
            indicator_description = indicator_description_element.text if indicator_description_element is not None else ''
            indicator_observable_element = indicator.find('.//indicator:Observable', namespaces=namespaces)
            indicator_observable = indicator_observable_element.attrib['idref'] if indicator_observable_element is not None else ''
            indicator_indicated_TTP_element = indicator.find('.//indicator:Indicated_TTP/stixCommon:TTP', namespaces=namespaces)
            indicator_indicated_TTP = indicator_indicated_TTP_element.attrib['idref'] if indicator_indicated_TTP_element is not None else ''
            indicator_producer_element = indicator.find('.//indicator:Producer//stixCommon:Name', namespaces=namespaces)
            indicator_producer = indicator_producer_element.text if indicator_producer_element is not None else ''
            indicator_produced_time_element = indicator.find('.//indicator:Producer//cyboxCommon:Produced_Time', namespaces=namespaces)
            indicator_produced_time = indicator_produced_time_element.text if indicator_produced_time_element is not None else 'NA'
            indicator_received_time_element = indicator.find('.//indicator:Producer//cyboxCommon:Received_Time', namespaces=namespaces)
            indicator_received_time = indicator_received_time_element.text if indicator_received_time_element is not None else 'NA'

            ws.append((indicator_id, indicator_ts, indicator_title, indicator_types, indicator_description,
                       indicator_observable, indicator_indicated_TTP, indicator_producer, indicator_produced_time,
                       indicator_received_time))

            # clean up
            indicator.clear()
            for ancestor in indicator.xpath('ancestor-or-self::*'):
                while ancestor.getprevious() is not None:
                    del ancestor.getparent()[0]

        context = etree.iterparse(feed, events=('end',), tag='{%s}TTPs' % namespaces['stix'])
        for event, ttp in context:
            ttp_id = ttp.xpath('.//stix:TTP', namespaces=namespaces)[0].attrib['id']
            ttp_ts = ttp.xpath('.//stix:TTP', namespaces=namespaces)[0].attrib['timestamp']
            ttp_title_element = ttp.find('.//ttp:Title', namespaces=namespaces)
            ttp_title = ttp_title_element.text if ttp_title_element is not None else 'NA'
            ttp_types_elements = ttp.findall('.//ttp:Type', namespaces=namespaces)
            ttp_types = ','.join([x.text for x in ttp_types_elements])
            ttp_name_element = ttp.findall('.//ttp:Name', namespaces=namespaces)
            ttp_names = ','.join([x.text for x in ttp_name_element]) if ttp_name_element is not None else 'NA'
            ttp_short_descr_element = ttp.find('.//ttp:Short_Description', namespaces=namespaces)
            ttp_short_descr = ttp_short_descr_element.text if ttp_short_descr_element is not None else ''
            ttp_descr_element = ttp.find('.//ttp:Description', namespaces=namespaces)
            ttp_descr = ttp_descr_element.text if ttp_descr_element is not None else ''

            ws2.append((ttp_id, ttp_ts, ttp_title, ttp_types, ttp_names, ttp_short_descr, ttp_descr))

            ttp.clear()
            for ancestor in ttp.xpath('ancestor-or-self::*'):
                while ancestor.getprevious() is not None:
                    del ancestor.getparent()[0]

        context = etree.iterparse(feed, events=('end',), tag='{%s}Observables' % namespaces['stix'])
        for event, observable in context:
            observ_id = observable.xpath('.//cybox:Observable', namespaces=namespaces)[0].attrib['id']
            observ_title_element = observable.find('.//cybox:Title', namespaces=namespaces)
            observ_title = observ_title_element.text if observ_title_element is not None else 'NA'
            observ_description_elements = observable.find('.//cybox:Description', namespaces=namespaces)
            observ_description = observ_description_elements.text if observ_description_elements is not None else 'NA'
            observ_properties_element = observable.find('.//cybox:Properties', namespaces=namespaces)
            observ_properties = ', '.join([': '.join([x.tag, x.text]) for x in observ_properties_element]) if observ_properties_element is not None else 'NA'

            if filter(lambda x: x != 'NA', [observ_title, observ_description, observ_properties]):
                ws3.append((observ_id, observ_title, observ_description, observ_properties))

            observable.clear()
            for ancestor in observable.xpath('ancestor-or-self::*'):
                while ancestor.getprevious() is not None:
                    del ancestor.getparent()[0]

        del context
        workbook.save(book_name)

        print 'Feed `{}` processed successfully.\nResulting file: `{}`'.format(feed, book_name)
        print '-' * 40


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Parses Hailataxii.com feeds and writes result into xls.')
    parser.add_argument("--feeds", "-f", help="filename(s) comma-separated", default='')
    args = parser.parse_args()
    feeds = args.feeds
    # feeds = 'results/dataForLast_7daysOnly_2015-11-24_04_29_46_247636.xml'

    feeds_arg = feeds.split(',') if feeds else []
    main(feeds=feeds_arg)
    print 'Job done.'
